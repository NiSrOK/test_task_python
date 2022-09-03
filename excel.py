import openpyxl
from openpyxl.writer.excel import save_workbook

FILE_NAME = 'exchange rates.xlsx'

def write_to_excel(info, currency_type):
    usd_columns = ['A', 'B', 'C']
    eur_columns = ['D', 'E', 'F']
    columns = eur_columns

    try:
        wb = openpyxl.load_workbook(FILE_NAME)
    except:
        wb = openpyxl.Workbook()
    ws = wb.active

    if currency_type == 'usd':
        columns = usd_columns

    for i in range(len(info)):
        for j in range(len(info[i])):
            if ',' in info[i][j]:
                info[i][j] = float(info[i][j].replace(',', '.'))
            else:
                info[i][j] = info[i][j][:6] + '20' + info[i][j][6:]

    ws[f'{columns[0]}1'] = 'Дата'
    ws[f'{columns[1]}1'] = 'Курс'
    ws[f'{columns[2]}1'] = 'Изменение'

    i = 2
    for day in info:
        ws[f'{columns[0]}{i}'] = day[0]
        ws[f'{columns[1]}{i}'] = day[1]
        ws[f'{columns[1]}{i}'].number_format = '#,##0.0000[$₽-419]'
        ws[f'{columns[2]}{i}'] = day[2]
        ws[f'{columns[2]}{i}'].number_format = '#,##0.00[$₽-419]'
        i += 1

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))    
    for col, value in dims.items():
        ws.column_dimensions[openpyxl.utils.cell.get_column_letter(col)].width = value * 1.23

    save_workbook(wb, FILE_NAME)      

def euro_to_dollar_ratio():
    ratio = []
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb.active
    for row in range(2, ws.max_row+1):
        ratio.append(ws.cell(row, 5).value/ws.cell(row, 2).value)
    for i in range(len(ratio)):
        ws[f'G{i+2}'] = ratio[i]
    
    save_workbook(wb, FILE_NAME)

def number_of_lines():
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb.active
    row_count = ws.max_row

    type_1 = ['строка', [1]]
    type_2 = ['строки', [2, 3, 4]]
    type_3 = ['строк', [5, 6, 7, 8, 9, 0]]

    all_types = [type_1, type_2, type_3]

    if row_count < 10:
        for type in all_types:
            if row_count in type[1]:
                return(f'В файле {row_count} {type[0]}.')
    elif row_count < 20:
        return(f'В файле {row_count} строк.')
    else:
        for type in all_types:
            if row_count % 10 in type[1]:
                return(f'В файле {row_count} {type[0]}.')  